import matplotlib.pyplot as plt
from matplotlib import cm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from itertools import product
import os

################################################################################
# Reading the data
################################################################################

"""
The DataFrame structure used for plotting:

1 - Index

The index should be 0-N logical indices (e.g., Strain, FLC) and 1 technical
index (_Source). _Source should uniquely identify a growth curve within a set of
logical indices.  It should be of the format:
Key1:Val1;Key2:Val2;[...]KeyN:ValN;

Note that while a set of logical indices must have one or more unique
_Source values, a _Source value need not be unique across different sets of
logical indices. This scenario happens in DataFrames that were averaged over.
Nevertheless, this should be treated as an edge case and avoided if possible.
For example, it can identify a specific well, or a plate/well combination.
Additionally, _Source must always be the last index. reorder_indices is a
helper function written to enforce that.

Downstream functions should declare which keys they expect, and it's up to the
user to make sure those keys exist and are correct. It is possible that not all
keys will exist in all DataFrames (e.g., the Well key can't remain after averaging).
"""

def read_plate_key(in_file_or_wb, sheet_name="Keys", converters=None, plate_type=96):
    """Read the plate keys from a given sheet in an Excel file.
    
    The key sheet should consist of one or more **tables** separated by at
    least one empty row, followed by an uppercase "STOP" (also separated by at
    least one empty row). Rows after the "STOP" are ignored. Everything should
    be aligned to the first column.
    
    The **table** format is as follows:
    The first cell of the first row should store the name of the key, e.g.
    "Strain" or "Media". The next row should be a N-by-12 grid, with every cell
    storing the key of the corresponding well (where the top-left cell
    corresponds to the A1 well). No empty rows are allowed between the key name
    and the key grid. N can be less than 8 - leading empty rows should contain
    at least one cell with '_' in it (which is always interpreted as "ignore 
    this cell"), otherwise all cells can be empty. Trailing empty rows are not
    necessary - the first empty row will indicate an end of the table.
    
    An example with a 3x3 layout (instead of 8x12) where the first row is to be
    ignored:
    
    +------+---+---+
    |Strain|   |   |
    +------+---+---+
    |   _  |   |   |
    +------+---+---+
    |   1  | 2 | 3 |
    +------+---+---+
    |   1  | 2 | 3 |
    +------+---+---+
    |      |   |   |
    +------+---+---+
    | Media|   |   |
    +------+---+---+
    |  _   |   |   |
    +------+---+---+
    | YPD  |YPD|YPD|
    +------+---+---+
    | SDC  |SDC|SDC|
    +------+---+---+
    |      |   |   |
    +------+---+---+
    | STOP |   |   |
    +------+---+---+
    
    Parameters
    ----------
    in_file_or_wb : str or openpyxl.Workbook
        The path to the Excel file holding the keys or said file already loaded
        into an openpyxl Workbook.
    sheet_name : str, default: ``"Keys"``
        The sheet name of the sheet that holds the keys.
    converters : dict of str to callable, optional
        A dictionary of converters for the key values. The dict keys are key
        names, while the dict values are callables that take the string value
        from the sheet and convert it to the correct value for further
        processing in Python.
    plate_type : number, optional
        The size of the plate - currently we support 96 and 384-well plates.
    
    Returns
    -------
    keys : list
        The keys parsed in the key sheet (list of strings). In the above example
        this would be ``('Strain', 'Media')``.
    key_values : dict of str to tuple
        A dictionary mapping well names to a tuple of
        key values (with the same length as the lengths of the returned `keys`).
        In the above example this would be ``{'A1': (1, 'YPD'), 'A2': (2, 'YPD'),
        'A3': (3, 'YPD'), 'B1': (1, 'SDC'), 'B2': (2, 'SDC'), 'B3': (3, 'SDC')}``.
    """
    
    assert plate_type in (96, 384)
    rows = list("ABCDEFGH" if plate_type == 96 else "ABCDEFGHIJKLMNOP")
    cols = list(range(1, 13 if plate_type == 96 else 25))
    if converters is None:
        converters = {}
    
    cells = [f"{r}{c}" for r, c in product(rows, cols)]
    result = {c: [] for c in cells}
    
    wb = _get_workbook(in_file_or_wb)
    sheet = wb[sheet_name]
    
    keys = []
    start_ix = stop_ix = -1 # The start and stop ixes of a single key table
    for row_ix, row in enumerate(sheet.iter_rows()):
        key = row[0].value
        if key is None or not str(key).strip():
            # This is an empty row
            if stop_ix < start_ix:
                # We've reached the end of a table
                stop_ix = row_ix - 1
                df_rows = rows[:stop_ix-start_ix+1]
                
                df = pd.read_excel(
                    wb,
                    sheet_name=sheet_name,
                    skiprows=start_ix,
                    header=None,
                    names=cols,
                    nrows=stop_ix-start_ix+1,
                    engine="openpyxl"
                )
                df["Rows"] = df_rows
                df.set_index("Rows", inplace=True)
                
                converter = converters.get(keys[-1], lambda x: x)
                for row, col in product(df_rows, cols):
                    if pd.isna(df.loc[row, col]) or "_" == df.loc[row, col]:
                        continue
                    result[f"{row}{col}"].append( converter(df.loc[row, col]) )
            else:
                # This is just an empty row, which we allow and simply skip
                continue
        elif stop_ix >= start_ix:
            # We're either expecting the start of a key table or the end of the table list
            key = str(key).strip()
            
            if key.upper() == "STOP":
                break
            
            keys.append(key)
            start_ix = row_ix+1
            
    for key in list(result.keys()):
        # A well label must exist in all keys for the well to be used:
        if not result[key] or len(result[key]) < len(keys):
            del result[key]
        else:
            result[key] = tuple(result[key])
        
    return keys, result

def _get_workbook(in_file_or_wb):
    if isinstance(in_file_or_wb, openpyxl.Workbook):
        return in_file_or_wb
    else:
        return load_workbook(filename=in_file_or_wb)

def read_od_sheet(
    in_file_or_wb, od_sheet_name, format, key_sheet_name,
    converters=None, plate_name=None, plate_type=96, over=70000
):
    """Parse plate growth curve data into a DataFrame.
    
    Reads the plate data along with the relevant keys (expected to be in the
    same workbook). The DataFrame will have a MultiIndex over the rows as
    parsed from the `key_sheet_name` sheet, with the last level called "_Source"
    and storing the plate and well information. The columns are "Time (s)", "OD"
    and "Temp" (temperature).
    
    Parameters
    ----------
    in_file_or_wb : str or openpyxl.Workbook
        The path to the Excel input file or said input file already loaded into
        an openpyxl Workbook.
    od_sheet_name : str
        The name of the sheet holding the plate data.
    format : str
        The format of the OD sheet. Currently we support: `f200`, `spark`,
        `spark stacker`.
    key_sheet_name : str
        The name of the sheet holding the key definitions.
    converters : dict, optional
        Converters for the key values, see the same parameter in the
        `read_plate_key` function.
    plate_name : str, optional
        The 'Plate' key for the _Source index. If `None`, will default to
        `od_sheet_name`.
    plate_type : number, optional
        The size of the plate - currently we support 96 and 384-well plates.
        Only relevant for the `spark stacker` format.
    over : number
        If the reading is outside of the dynamic range, the Stacker will record
        "OVER". This value will replace it. Only relevant for the
        `spark stacker` format.

    Returns
    -------
    pandas.DataFrame
        The parsed OD measurements for the plate.
    """

    wb = _get_workbook(in_file_or_wb)
    keys, plate_map = read_plate_key(wb, key_sheet_name, converters, plate_type)
    return read_od_sheet_with_plate_map(wb, od_sheet_name, keys, plate_map, format, plate_name, plate_type, over)
    
def read_od_sheet_with_plate_map(
    in_file_or_wb, od_sheet_name, keys, plate_map, format,
    plate_name=None, plate_type=96, over=70000
):
    if plate_name is None:
        plate_name = od_sheet_name

    sheet = _get_workbook(in_file_or_wb)[od_sheet_name]
    new_data = { "OD": [], "Time (s)": [], "Temp": [] }
    index = []

    if format == "f200":
        read_od_sheet_f200(sheet, plate_map, plate_name, new_data, index)
    elif format == "spark":
        read_od_sheet_spark(sheet, plate_map, plate_name, new_data, index)
    elif format == "spark stacker":
        read_od_sheet_spark_stacker(sheet, plate_map, plate_name, plate_type, new_data, index, over)
    else:
        assert False, "OD sheet format is not recognized"

    return pd.DataFrame(
        new_data,
        index=pd.MultiIndex.from_tuples(index, names=list(keys)+["_Source"])
    ).sort_index()

def read_od_sheet_f200(sheet, plate_map, plate_name, data_dict, index):
    for row_ix, row in enumerate(sheet.iter_rows()):
        if "Time [s]" in str(row[0].value):
            break
    
    df = pd.read_excel(
        sheet.parent, engine="openpyxl",
        sheet_name=sheet.title,
        skiprows=row_ix,
        header=None,
    ).dropna(axis=0, how="all").dropna(axis=1, how="all").iloc[:-1,:]
    
    time_series = list(df.iloc[0,1:])
    temp_series = list(df.iloc[1,1:])

    for well_ix, well in enumerate(df.iloc[2:,0]):
        data_dict["OD"] += list(df.iloc[well_ix+2,1:])
        data_dict["Time (s)"] += time_series
        data_dict["Temp"] += temp_series
        
        index += [plate_map[well] + (f"Plate:{plate_name};Well:{well};",)] * len(time_series)

def read_od_sheet_spark_stacker(sheet, plate_map, plate_name, plate_type, data_dict, index, over):
    assert plate_type in (96, 384), "Plate type can only be 96 or 384"
    row_num = 8 if plate_type == 96 else 8*2
    col_num = 12 if plate_type == 96 else 12*2
    
    table_rows_to_read = 0
    time = temp = None
    for row_ix, row in enumerate(sheet.iter_rows()):
        cell_value = row[0].value
        if table_rows_to_read > 0:
            for i in range(1, col_num+1):
                measurement_str = row[i].value
                well = f"{cell_value}{i}"
                if well not in plate_map:
                    continue
                
                # The Spark Stacker can give an "OVER" measurement if we're
                # outside of the dynamic range:
                data_dict["OD"].append(float(measurement_str) if measurement_str != "OVER" else over)
                data_dict["Time (s)"].append(time)
                data_dict["Temp"].append(temp)
                
                index.append(plate_map[well] + (f"Plate:{plate_name};Well:{well};",))

            table_rows_to_read -= 1
        if cell_value == "End Time":
            break
        elif cell_value is None or cell_value.strip() == "":
            continue
        elif cell_value == "Time [s]":
            time = row[1].value
        elif cell_value == "Temp. [°C]":
            temp = row[1].value
        elif cell_value == "<>":
            # These will be read starting with the next iteration
            table_rows_to_read = row_num

def read_od_sheet_spark(sheet, plate_map, plate_name, data_dict, index):
    for row_ix, row in enumerate(sheet.iter_rows()):
        cell_value = row[0].value
        if cell_value == "STOP" or cell_value == "End Time":
            break
        elif cell_value in ("Temp. [°C]", "Time [ms]") or cell_value in plate_map:
            # First, read the row until the end:
            row_values = []#[int(c.value) for c in row[1:]]
            for next_cell in row[1:]:
                next_value = next_cell.value
                if next_value is None or str(next_value).strip() == "":
                    break
                row_values.append(next_value)
            
            # Then decide what to do with it:
            if cell_value == "Temp. [°C]":
                temp_row = row_values
            elif cell_value in plate_map:
                # Will be skipped if cell_value is not in plate map!
                well = cell_value
                ods_row = row_values
            else:
                assert cell_value == "Time [ms]"

                # The well could be empty, in which case we ignore it:
                if len(row_values) == 0:
                    continue

                assert len(ods_row) == len(row_values) == len(temp_row)

                data_dict["OD"] += ods_row
                data_dict["Time (s)"] += [int(t)/1000 for t in row_values]
                data_dict["Temp"] += temp_row

                index += [plate_map[well] + (f"Plate:{plate_name};Well:{well};",)] * len(ods_row)

################################################################################
# Reading utilities
################################################################################

def read_384_htl(fname, plate_start, key_start, num_of_plates=9, extra_source=None):
    wb = load_workbook(fname)

    dfs_to_concat = []
    
    for i in range(num_of_plates):
        
        keys, plate_map = read_plate_key(wb, f"Plate {key_start+i} keys", plate_type=384)
        df = read_od_sheet_with_plate_map(wb, f"Plate {plate_start+i}", keys, plate_map, "spark stacker",
                                          plate_type=384, plate_name=f"Plate {plate_start+i:02}",)
        dfs_to_concat.append(df)
    
    result = pd.concat(dfs_to_concat).sort_index()

    if extra_source:
        result = result.reset_index("_Source")
        result["_Source"] = extra_source + result["_Source"]
        result = result.set_index("_Source", append=True)

    return result

# TODO: refactor read_384_htl into read_384_htl_with_origin
# The origin is useful when comparing wells between different experiments that
# came from the same thawed spot.
def read_384_htl_with_origin(fname, plate_start, key_start, num_of_plates=9, extra_source=None):
    wb = load_workbook(fname)

    dfs_to_concat = []

    from string import ascii_uppercase
    from math import ceil
    from growth_curves import _parse_source
    for i in range(num_of_plates):
        
        keys, plate_map = read_plate_key(wb, f"Plate {key_start+i} keys", plate_type=384)
        df = read_od_sheet_with_plate_map(wb, f"Plate {plate_start+i}", keys, plate_map, "spark stacker",
                                          plate_type=384, plate_name=f"Plate {plate_start+i:02}",)

        wells = [p["Well"] for p in _parse_source(df)]
        row_coords = [ceil((ascii_uppercase.index(w[0])+1) / 2) for w in wells]
        unique_row_num = len(set(row_coords))
        col_coords = [ceil(int(w[1:]) / 2) for w in wells]
        # TODO: assumes the growth curve plates come in consecutive batches of 3 per origin plate.
        # 24: cols in a row
        # 1000: arbtirary decision, as long as it's >=384 it's good. 1000 was chosen for ease 
        # of debugging (since the origins shouldn't exceed 3 
        df["Origin"] = [c + (r-1) * 24 + 1000*(i//3) for r, c in zip(row_coords, col_coords)]
        df = reorder_indices(df.set_index("Origin", append=True))
        
        dfs_to_concat.append(df)
    
    result = pd.concat(dfs_to_concat).sort_index()

    if extra_source:
        result = result.reset_index("_Source")
        result["_Source"] = extra_source + result["_Source"]
        result = result.set_index("_Source", append=True)

    return result

################################################################################
# Plotting the data
################################################################################

def plot_ods(
    df,
    x_index=None, y_index=None, x_title=None, y_title=None,
    x_index_grid=None,
    x_col="Time (s)", y_col="OD",
    x_index_key=None, y_index_key=None,
    mean_indices=None,
    std_dev=None, # None, "bar", "area" -- must be used with mean_indices!
    style_func=None,
    legend="last col",
    title_all_axes=False,
    cmap="viridis",
    dpi=150,
    alpha=1,
    figsize_x_scale=1,
    figsize_y_scale=1,
    ax_func=None,
    sharey=True,
    legend_ncol=1,
):
    """Plot the ODs from a DataFrame returned by `read_experiment`.
    
    Parameters
    ----------
    df : pandas.DataFrame
        The DataFrame with the growth curves to plot.
    x_index : str, optional
        The name of the level in `df`'s index to plot along the X axis
        (columns) of the figure. The values of this level will be used as the
        titles for the first row of Axes in the figure.
    y_index : str, optional
        The name of the level in `df`'s index to plot along the Y axis (rows)
        of the figure. The values of this level will be used as the title of
        the y axis for the first column of the Axes in the figure.
    x_index_grid : sequence of sequences, optional
        If no `y_index` is given, this allows to break the `x_index` across multiple
        rows. Defines the logical layout of the figure according to the `x_index`
        labels. For example, two rows of three drug concentrations in each row:
        `[[0, 8, 16], [32, 64, 128]]`.
    x_col : str, optional
        The column holding the x axis values for plotting. Default is 'Time (s)',
        which will be converted to hours. No conversions will be done for other
        values.
    y_col : str, optional
        The column holding the y axis values for plotting. Default is 'OD'.
    x_title : str, optional
        If x_index is not specified, this will be used as the title of the
        first Axes in the figure.
    y_title: str, optional
        If y_index is not specified, this will be used as the title of the y
        axis of the first Axes in the figure.
    x_index_key : callable, optional
        A sorting key function used to order the labels of the x axis for plotting.
    y_index_key : callabble, optional
        As `x_index_key`, but for the y axis labels.
    mean_indices : sequence of str, optional
        The index levels which will be used to group the growth curves for
        averaging. For example, if ``("Strain", "Media")`` is passed, every unique
        strain-media combination will be grouped, and all remaining sub-levels
        will be averaged over to yield a single growth curve, which will then
        be plotted. If different experiments within the group have different time stamps,
        the "missing" measurements will be interpolated before averaging.
    std_dev : {None, 'bar', 'area'}
        If `mean_indices` is specified, this will set the method for displaying
        the standard deviation:
            
        - **bar** - show the standard deviation as whiskers.
        - **area** - show the standard deviation as a semi-transparent area.
    style_func : callable, optional
        A callable that accepts three parameters ``(x_label, y_label, ix)`` where ix is
        the index of the curve to draw. It should return a dict of kwargs for
        the `Axes.plot` method.
    legend : {"last col", "every axes", "none"}
        - ``"last col"`` will add a legend at the end of each row.
        - ``"every axes"`` will add a legend to every axes.
        - ``"none"`` will not show any legends.
    title_all_axes : bool, default=False
        If True, will set the title for all Axes (not just the top row).
    cmap : str or `matplotlib.colors.Colormap`, default='viridis'
        The colormap to use for the growth curves.
    ax_func : callable, optional
        A callable that takes three parameters: `ax, x_label, y_label`. `ax`
        is an Axes object, and `x_label` and `y_label` are its labels (indices).
        Used to plot extra things on individual Axes, e.g. 24-hour vertical lines.
    sharey : bool or str, optional
        Will be passed on to `plt.subplots`. It may be useful to only share the
        y axis across rows (pass `"rows"`) or have no sharing at all (pass `False`).
    
    Returns
    -------
    matplotlib.figure.Figure
        The figure with the plots.
    """
    
    if std_dev is not None and mean_indices is None:
        raise ValueError("std_dev is not None but mean_indices was not given!")
    
    if isinstance(cmap, str):
        cmap = cm.get_cmap(cmap)
    
    if mean_indices is not None:
        # NB: mean_indices can be an empty sequence!
        df = avg_over_ixs(df, mean_indices)
    else:
        # Work on a copy of the df, in case we'll need to modify it:
        df = df.copy()
    
    # If the x/y indices are not specified, create singleton "dummy" levels to
    # allow for the subsequent code to be agnostic of these effects.
    if x_index is None:
        x_index = "_dummy_x"
        df[x_index] = ""
        df.set_index(x_index, append=True, drop=True, inplace=True)
    if y_index is None:
        y_index = "_dummy_y"
        if x_index_grid:
            row_ixs = []
            for ix in df.index.get_level_values(x_index):
                for row_ix, row_vals in enumerate(x_index_grid):
                    if ix in row_vals:
                        row_ixs.append(row_ix)
                        break
            df[y_index] = row_ixs
        else:
            df[y_index] = ""
        df.set_index(y_index, append=True, drop=True, inplace=True)
    # In case any indices were added after _Source:
    df = reorder_indices(df)
    
    x_labels = df.index.get_level_values(x_index).unique()
    y_labels = df.index.get_level_values(y_index).unique()
    if x_index_grid:
        y_labels = sorted(y_labels)

    if x_index_key is not None:
        x_labels = list(sorted(x_labels, key=x_index_key))
    if y_index_key is not None:
        y_labels = list(sorted(y_labels, key=y_index_key))
    
    rows = len(y_labels)
    cols = len(x_labels)
    if x_index_grid:
        cols = max(len(r) for r in x_index_grid)
    
    fig, axs = plt.subplots(
        rows, cols, figsize=(cols*3*figsize_x_scale, rows*2*figsize_y_scale),
        sharex=True, sharey=sharey, dpi=dpi,
        squeeze=False, layout="compressed",
    )
    
    if x_index_grid:
        for row_ix in range(len(x_index_grid)):
            for ix, label in enumerate(x_index_grid[row_ix]):
                axs[row_ix][ix].set_title(x_title or label)
    else:
        for ix, label in enumerate(x_labels):
            axs[0][ix].set_title(x_title or label)
        
        for ix, label in enumerate(y_labels):
            axs[ix][0].set_ylabel(y_title or label)
        
    for row_ix, y_label in enumerate(y_labels):
        for col_ix, x_label in enumerate(x_labels):
            if x_index_grid:
                if x_label not in x_index_grid[y_label]:
                    continue
                col_ix = x_index_grid[y_label].index(x_label)

            ax = axs[row_ix][col_ix]
            ax.xaxis.set_tick_params(which='both', labelbottom=True)
            
            # Not all indexes are guaranteed to be in every axes. Skip them if
            # they don't exist.
            if x_label not in df.index.get_level_values(x_index) or \
                y_label not in df.index.get_level_values(y_index):
                continue
            ax_df = df.xs((x_label, y_label), level=(x_index, y_index))
            
            # The same condition can come from several sources (e.g., wells),
            # and this gets special treatment in the condition loop. To set up
            # the condition loop, we need to get all of the condition indexes,
            # but without the source information.
            # A potential problem is that if _Source is the last index (i.e.,
            # there's only one curve per Axes), we can't drop it, so we have
            # to test for it.
            if ax_df.index.nlevels > 1:
                condition_ixs = ax_df.index.droplevel("_Source").unique()
            else:
                condition_ixs = ax_df.index.unique()
                
            for con_ix_count, con_ix in enumerate(condition_ixs):
                con_df = ax_df.loc[con_ix]
                
                # TODO: colors break when an index is missing from one of the axes
                if len(condition_ixs) == 1:
                    color = "black"
                else:
                    color = cmap(1.0*(con_ix_count/(len(condition_ixs)-1)))
                
                if mean_indices is not None:
                    con_df = con_df.copy() # To prevent SettingWithCopyWarning
                    if std_dev is not None:
                        std_dev_df = con_df["OD std"].reset_index()
                        std_dev_df["OD"] = std_dev_df["OD std"]
                    con_df["OD"] = con_df["OD mean"]
                    
                wells = con_df.index.get_level_values("_Source").unique()
                
                for well_ix, well in enumerate(wells):
                    well_df = con_df.loc[well]
                    style = {
                        "label": con_ix,
                        "color": color,
                        "alpha": alpha,
                    }
                    if style_func is not None:
                        user_style = style_func(x_label, y_label, con_ix)
                        style.update(user_style)
                    if well_ix > 0:
                        # This will hide this entry from the legend:
                        style["label"] = f"_{style['label']}"
                    
                    if x_col == "Time (s)":
                        xs = well_df["Time (s)"] / 60 / 60
                    else:
                        xs = well_df[x_col]
                    ys = well_df[y_col]
                    if not std_dev or std_dev == "area":
                        ax.plot(xs, ys, **style)
                        if std_dev == "area":
                            # NB: reset_index() is required here because the index
                            # of ys is a running number while the index of
                            # std_dev_data is an empty string (due to how it's created)
                            # NB: .tolist() is needed for some versions of matplotlib
                            # which throw a weird error if a float64 Series is passed.
                            ax.fill_between(
                                xs.tolist(),
                                (ys.reset_index(drop=True) - std_dev_df["OD"]).tolist(),
                                (ys.reset_index(drop=True) + std_dev_df["OD"]).tolist(),
                                color=style["color"], alpha=0.2
                            )
                    elif std_dev == "bar":
                        ax.errorbar(
                            xs, ys, yerr=std_dev_df["OD"], errorevery=4,
                            **style
                        )
                
            if ax_func is not None:
                ax_func(ax, x_label, y_label)
            
            if len(condition_ixs) == 1 and title_all_axes:
                prev_title = ax.get_title()
                if prev_title:
                    prev_title += "\n"
                ax.set_title(prev_title + str(condition_ixs[0]))
                
            if legend == "every axes":
                legend_without_duplicate_labels(
                    ax,
                    fontsize=8, loc='upper left',
                    ncol=legend_ncol
                )

    for ax_row in axs:
        for ax in ax_row:
            if not ax.lines and not ax.collections:
                ax.set_axis_off()

    if legend == "last col":
        for ax_row in axs:
            # TODO: this gives a warning if some of the Axes on the row have no
            # values, and in any case will only display the legend for the
            # rightmost Axes.
            legend_without_duplicate_labels(
                ax_row[-1],
                bbox_to_anchor=(1.05, 1),
                loc='upper left', borderaxespad=0., fontsize=6,
                ncol=legend_ncol
            )
    
    return fig

def avg_over_ixs(df_in, avg_levels, x_col="Time (s)", y_col="OD", interpolation_method="from_derivatives"):
    """Return a DataFrame that averages all levels over a given set of levels.
    
    If the timestamps differ between any measurement series, measurements will
    be interpolated for the missing timestamps.
    
    The new DataFrame will have `avg_levels` as the row MultiIndex and three
    columns: `x_col`, `y_col mean` and `y_col std` (for standard deviation).
    
    Parameters
    ----------
    df_in : pandas.DataFrame
        The input DataFrame.
    avg_levels : sequence of str
        The index levels to group by for averaging purposes.
    x_col : str, default: "Time (s)"
        The independent variable column name.
    y_col : str, default: "OD"
        The dependent variable column name.
    interpolation_method : str, default: "from_derivatives"
        The interpolation method.
    
    Returns
    -------
    pandas.DataFrame
        A new DataFrame, with `avg_levels` as the new MultiIndex, and three 
    """

    df_in = reorder_indices(df_in, avg_levels) # Just in case the user hasn't.
    dfs_to_concat = []
    for exp_ix in df_in.index.droplevel([i for i in df_in.index.names if i not in avg_levels]).unique():
        # If avg_levels is only 1 level, exp_ix must be stored into a tuple
        # for df_in.xs to work:
        if len(avg_levels) == 1:
            exp_ix = (exp_ix,)
        exp_df = df_in.xs(exp_ix, level=avg_levels)[[x_col, y_col]]
        
        avg_dfs = []
        # No assignments are made here, but pandas still gives a warning, hence
        # we suppress it.
        pd.set_option('mode.chained_assignment', None)
        for avg_ix in exp_df.index.unique():
            avg_dfs.append( exp_df.loc[avg_ix].set_index(x_col).sort_index() )
        pd.set_option('mode.chained_assignment','warn')
        
        joined_df = avg_dfs[0]
        for col_suffix, df in enumerate(avg_dfs[1:]):
            joined_df = joined_df.join(df, how="outer", sort=True, rsuffix=str(col_suffix))
            
        for col_name in joined_df.columns:
            joined_df.loc[:,col_name] = joined_df.loc[:,col_name].interpolate(method=interpolation_method)
            
        joined_df.dropna(inplace=True) # For the extreme NaNs that don't get interpolated
        
        mean_s = joined_df.mean(axis=1)
        std_s = joined_df.std(axis=1).fillna(0) # Std. dev. will be NaN if there's only one value to average
        joined_df[f"{y_col} mean"] = mean_s
        joined_df[f"{y_col} std"] = std_s
        
        joined_df = pd.DataFrame({f"{y_col} mean": mean_s, f"{y_col} std": std_s}).reset_index()
        # TODO: is there a better way to set a single multi-index tuple into a dataframe?
        for level_name, level_value in zip(avg_levels, exp_ix):
            joined_df[level_name] = level_value

        # NB: we set the number of averaged DataFrames as the _Source for ease of
        # future debugging. However, there is an assumption here that averaged
        # DataFrames will not be further concatenated, otherwise their _Source
        # indices will clash, which might lead to issues if their logical indices
        # also clash.
        joined_df["_Source"] = f"Avg:{len(avg_dfs)}"
        
        joined_df.set_index(list(avg_levels) + ["_Source"], inplace=True)
        dfs_to_concat.append(joined_df)
        
    return pd.concat(dfs_to_concat)

# Example of a style function:
def style_func(x_label, y_label, ax_index):
    return {"label": "", "color": "", "linestyle": ""}

def _parse_source(df):
    """Return a list of dicts which maps the Key:Val pairs in the _Source index.
    """
    
    result = []

    for ix in df.index.get_level_values("_Source"):
        row_kvs = {}
        for kv_pair in ix.split(';'):
            if not kv_pair:
                continue
            k, v = kv_pair.split(":")
            row_kvs[k] = v
        result.append(row_kvs)

    return result

def reorder_indices(df, head=None):
    """Return a new DataFrame with a new order of the indices, such that the
    _Source index is last. Optionally, the first indices can be forced by the
    `head` parameter.
    """
    if head is None:
        head = []
    ix_names = head + \
        [n for n in df.index.names if n not in head and n != "_Source"] + \
        ["_Source"]
    return df.reorder_levels(ix_names)

def add_row_col(df):
    # Assumes the df comes from a single plate!
    result = df.copy()
    well_values = [ix["Well"] for ix in _parse_source(df)]
    result["Row"] = [w[0] for w in well_values]
    result["Column"] = [w[1:] for w in well_values]
    result.set_index("Row", inplace=True, append=True)
    result.set_index("Column", inplace=True, append=True)
    
    return result

def plot_plate(df):
    """NB: assumes the '_Source' index has a 'Well' key!"""
    return plot_ods(
        reorder_indices(add_row_col(df).sort_index(level="Column").sort_index(level="Row")),
        x_index="Column", y_index="Row", legend="every axes",
        x_index_key=int
    )

# Adapted from https://stackoverflow.com/a/56253636
def legend_without_duplicate_labels(ax, **kws):
    handles, labels = ax.get_legend_handles_labels()
    unique = [(h, l) for i, (h, l) in enumerate(zip(handles, labels)) if l not in labels[:i]]
    ax.legend(*zip(*unique), **kws)
    
################################################################################
# Utilities
################################################################################

def xss(df, keys, levels, drop_singleton_levels=False):
    level_ixs = [df.index.names.index(l) for l in levels]
    index_dict = {l_ix: vs for (l_ix, vs) in zip(level_ixs, keys)}
    
    def test_index(index):
        for index_ix, index_value in enumerate(index):
            if index_ix in index_dict:
                if index_value not in index_dict[index_ix]:
                    return False
        
        return True
    
    result = df.loc[df.index.map(test_index)]
    
    if drop_singleton_levels:
        for level_ix in range(result.index.nlevels-1, -1, -1):
            if len(result.index.get_level_values(level_ix).unique()) <= 1:
                result = result.droplevel(level_ix)
    
    return result    

################################################################################
# Configurable plotting from Excel
################################################################################

_FIG_LABEL = "FIGURE"
_STOP_LABEL = "STOP"
_GLOBAL_LABEL = "GLOBAL"
_FORCE_LABEL = "FORCE"

# TODO: document the configuration format for the configuration file.
def make_plots_from_excel(df, excel_file, output_folder, sheet_name="Sheet1", draw_marked_only=False):
    """Generate plots from an Excel configuration file.
    
    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame in the format returned by `read_experiment`-like functions.
    excel_file : str
        Path to the configuration Excel file.
    output_folder : str
        Path to the folder where the generated plots will be stored.
    sheet_name : str, optional
        The sheet name in the configuration file to read. Default: `Sheet1`.
    draw_marked_only : bool, optional
        Specifies if only figures marked with a "FORCE" flag should be generated.
        Default: `False` (generates all figures).
    """
    
    global_params = {}

    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    row_ix = 1
    while row_ix <= sheet.max_row:
        lead_value = sheet.cell(row_ix, 1).value
        if not lead_value:
            row_ix += 1
            continue

        lead_value = lead_value.upper()
        if lead_value == _STOP_LABEL:
            break
        elif lead_value == _FIG_LABEL:
            fig_name = sheet.cell(row_ix, 2).value
            (fig_df, params), row_ix = _parse_figure(sheet, row_ix+1)
            if draw_marked_only and _FORCE_LABEL not in params:
                continue
            if _FORCE_LABEL in params:
                del params[_FORCE_LABEL]
            
            kws = global_params.copy()
            kws.update(params)
            fig = plot_ods(
                df.loc[fig_df.index],
                style_func=_style_func_by_df(fig_df),
                **kws
            )
            fig.savefig(os.path.join(output_folder, f"{fig_name}.png"))
            plt.close(fig)
        elif lead_value == _GLOBAL_LABEL:
            global_params, row_ix = _parse_params(sheet, row_ix+1)
        else:
            row_ix += 1

# Returns (result, next_ix)
def _parse_params(sheet, row_ix):
    result = {}

    while row_ix <= sheet.max_row:
        lead_value = sheet.cell(row_ix, 1).value
        if not lead_value:
            row_ix += 1
            continue

        if lead_value.upper() in (_FIG_LABEL, _STOP_LABEL, _GLOBAL_LABEL):
            break

        param_name = lead_value
        if param_name != "x_index_grid":
            param_value = sheet.cell(row_ix, 2).value
            row_ix += 1
        else:
            grid = param_value = []
            row_ix += 1
            while True:
                lead_value = sheet.cell(row_ix, 1).value
                if lead_value is None:
                    break

                row = [lead_value]
                col_ix = 2
                while True:
                    value = sheet.cell(row_ix, col_ix).value
                    if not value:
                        break
                    row.append(value)
                    col_ix += 1
                row = [None if i == "_" else i for i in row]
                grid.append(row)
                row_ix += 1

        result[param_name] = param_value

    return result, row_ix

def _parse_figure(sheet, row_ix):
    start_row_ix = row_ix
    while sheet.cell(row_ix, 1).value is not None and row_ix <= sheet.max_row:
        row_ix += 1
    
    df = pd.read_excel(
        pd.ExcelFile(sheet.parent, engine="openpyxl"),
        sheet_name=sheet.title,
        skiprows=start_row_ix-1,
        nrows=row_ix-start_row_ix-1,
        index_col=0,
    )

    params, row_ix = _parse_params(sheet, row_ix)

    return (df, params), row_ix

def _style_func_by_df(df):
    def _style_func(x_label, y_label, ix):
        return df.loc[ix].dropna().to_dict()
    return _style_func
