from typing import Optional, Dict, Any, Literal
import pandas as pd
from growth_curves import reorder_indices, xss
from itertools import product
import numpy as np

def get_curveball_fitness(df, with_tqdm=False, models=None):
    import curveball
    import curveball.models as cb_models

    ix_wrapper = lambda i: i
    if with_tqdm:
        from tqdm.notebook import tqdm
        ix_wrapper = tqdm

    df = df.copy()
    df["Time"] = df["Time (s)"] / 60 / 60

    all_models = {
        "ix": [],
        "lag": [],
        "min doubling time": [],
        "max OD": [],
        "models": [],
        "model name": [],
    }

    for ix in ix_wrapper(df.index.unique()):
        all_models["ix"].append(ix)
       
        model_fits = curveball.models.fit_model(df.loc[ix], PLOT=False, PRINT=False, models=models)
        all_models["models"].append(model_fits)
        best_model = model_fits[0]

        all_models["lag"].append(cb_models.find_lag(best_model))
        all_models["min doubling time"].append(cb_models.find_min_doubling_time(best_model))
        all_models["max OD"].append(df.loc[ix, "OD"].max())
        all_models["model name"].append(best_model.model.name)

    return pd.DataFrame(
        all_models,
        pd.MultiIndex.from_tuples(all_models["ix"], names=df.index.names)
    ).drop(["ix"], axis=1)

def move_columns_to_axis(df, columns, level_name, value_col_name):
    dfs_to_concat = []
    for col in columns:
        other_cols = [c for c in columns if c != col]
        subdf = df.loc[:,[c for c in df.columns if c not in other_cols]].\
            rename({col: value_col_name}, axis=1)
        subdf[level_name] = col
        subdf.set_index(level_name, inplace=True, append=True)

        dfs_to_concat.append(subdf)

    return reorder_indices(pd.concat(dfs_to_concat).sort_index())

# TODO: Older fitness computation + glucose calculations for Rhodamine 6G experiments.
# See if still needed and/or if should be improved upon.
"""
Every well gets a nominal fitness computation using a function (e.g., AUC).

Then we need to normalize. The normalization function can take the index of
a well, and return the index relative to which it should be normalized.

Optionally, we can avg out the wells.
"""

def auc_fitness(data):
    from numpy import trapz
    
    return trapz(data["OD"], data["Time (s)"])

def get_nominal_fitness(df, fit_func):
    # AUC caclulations are more interpretable if the baseline is zero, and not
    # some high arbitrary number.
    # TODO: allow normalizing against a sterile control.
    df = df.copy()
    df["OD"] = df["OD"] - df["OD"].min()
    data = {"Nominal": []}
    index = []
    
    for row_ix in df.index.unique():
        nominal_fit = fit_func(df.loc[row_ix])
        data["Nominal"].append(nominal_fit)
        index.append(row_ix)
    
    return pd.DataFrame(data, pd.MultiIndex.from_tuples(index, names=df.index.names)).sort_index()
    
def normalize_nominal_fit_df(nominal_fit_df, norm_func):
    # Average out the fitnesses:
    # nominal_mean_df = nominal_fit_df.groupby(level=nominal_fit_df.index.names[:-1]).mean()
    # result = nominal_mean_df.rename({"Nominal": "Normalized"}, axis="columns")
    result = nominal_fit_df.groupby(level=nominal_fit_df.index.names[:-1]).mean().rename({"Nominal": "Nominal mean"}, axis="columns")
    result["Normalized"] = -1
    
    for ix in result.index.unique():
        result.loc[ix, "Normalized"] = result.loc[ix]["Nominal mean"] / result.loc[norm_func(ix)]["Nominal mean"]
    
    return result

def norm_by_od(df, od_df):
    """Normalize the values in a given DataFrame by ODs from another DataFrame.
    
    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame to normalize in the format returned by `read_experiment`-like
        functions.
    od_df : pandas.DataFrame
        DataFrame of ODs. Rows should be indexed A-to-F, and columns 1-to-12.
        `df` will be normalized by using the 'Well' index level.
        
    Returns
    -------
    pandas.DataFrame
        A normalized copy of `df`.
    """
    
    result = df.copy()

    pre_well_ixs = (slice(None),) * (result.index.nlevels-1)
    for r, c in product(od_df.index, od_df.columns):
        ix = pre_well_ixs + (f"{r}{c}",)
        
        if not pd.isna(od_df.loc[r, c]) and \
            ix[-1] in df.index.get_level_values("Well"):
            result.loc[ix, "OD"] /= od_df.loc[r, c]
    
    return result

def normalize_glucose(df):
    """For R6G efflux experiments - normalize the readings from +Glu to -Glu
    condition.
    
    Assumes that every -/+ Glu condition has exactly two wells - so if there
    are any repeats, they either need to be separated by a key in the index
    or handled separately.
    
    Only the OD and Time (s) columns will be kept. The "Well" index will be
    chosen arbitrarily from the two -/+ Glu wells.
    
    Parameters
    ----------
    df : pandas.DataFrame
    
    Returns
    -------
    pandas.DataFrame
        A normalized version of `df`. The "Glucose" index level will not be
        part of the index.
    """
    
    diff_dict = {"OD": [], "Time (s)": []}
    index = []
    
    no_gluc_df = xss(df, [[0]], ["Glucose"], drop_singleton_levels=True)
    add_gluc_df = xss(df, [[1]], ["Glucose"], drop_singleton_levels=True)

    for ix in df.index.droplevel(["Glucose", "Well"]).unique():
        without_gluc = no_gluc_df.loc[ix].sort_values("Time (s)")
        with_gluc = add_gluc_df.loc[ix].sort_values("Time (s)")

        well = with_gluc.index[0]

        delta = with_gluc["OD"].reset_index(drop=True) - without_gluc["OD"].reset_index(drop=True)

        diff_dict["OD"] += list(delta)
        diff_dict["Time (s)"] += list(without_gluc["Time (s)"])

        index += [ix + (well,)]*len(delta)

    index = pd.MultiIndex.from_tuples(index, names=df.index.droplevel(["Glucose"]).names)
    return pd.DataFrame(diff_dict, index=index).sort_index()

################################################################################
# Utilities
################################################################################

# Sanitizing the growth curves before fitness estimations is strongly recommended.
# Artifacts in reading the OD can really mess with naive estimators.

def sanitize_growth_curves(df: pd.DataFrame) -> pd.DataFrame:
    """Remove initial artifacts from growth curves.

    Applies a per-curve filter that removes an initial monotonically decreasing
    segment in OD values, which may arise from measurement artifacts.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing growth curves. Must include an "OD" column
        and a MultiIndex identifying individual curves.

    Returns
    -------
    pandas.DataFrame
        DataFrame with the same structure as input, but with initial decreasing
        segments removed from each growth curve.

    Notes
    -----
    - The filtering is applied independently to each index group.
    - The current implementation removes the longest prefix where OD decreases.
    """

    # TODO: ChatGPT added the following note:
    # - This may remove entire curves if they are strictly decreasing.
    # Is this actually the case, and if so, should we do anything about it?

    result = df.copy()
    
    # For some reason, sometimes a few of the first readings have a high OD, which then collapses back to the baseline.
    # We should filter them out:
    ### This version looks for the first cutoff after which the ODs are monotonically rising.
    ### This can cut most of the 'EMPTY' curves, which can create problems for the fitness estimations.
    ### We could filter them out, but it's best to think of another method of filtering.
    # def filter_spikes(df):
    #     min_od = df["OD"].min()
    #     ix_cutoff = 0
    #     for od in df["OD"]:
    #         # if od > min_od * 1.5:
    #         if od > min_od:
    #             ix_cutoff += 1
    #         else:
    #             break
    #     if ix_cutoff == 0:
    #         return df
    #     else:
    #         # if ix_cutoff > 2:
    #         #     print(df.index[0], ix_cutoff)
    #         return df.iloc[ix_cutoff:]

    # This method cuts the first monotonically decreasing portion:
    def filter_spikes(df: pd.DataFrame) -> pd.DataFrame:
        diff = df["OD"].iloc[:-1] > df["OD"].iloc[1:]
        ix_cutoff = diff.argmin()
        return df.iloc[ix_cutoff:]

    # TODO For some reason, an index level is added which is just the tuple of all other indices?!
    result = result.groupby(result.index).apply(filter_spikes).reset_index(level=0, drop=True)
    
    return result

################################################################################
# Model-free fitness
################################################################################

# Nominal fitness 

def get_model_free_fitness_single(
    df: pd.DataFrame,
    lag_threshold: float = 2,
    plot: bool = False
) -> pd.Series:
    """Compute model-free fitness metrics for a single growth curve.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame representing a single growth curve. Must contain columns
        "OD" and "Time (s)".
    lag_threshold : float, default: 2
        Threshold multiplier relative to the minimum OD used to estimate lag.
    plot : bool, default: False
        If True, generates diagnostic plots of OD and derivatives.

    Returns
    -------
    pandas.Series
        Series containing the following metrics:
        - "Max OD"
        - "Lag (s)"
        - "Lag (hr)"
        - "Max slope"
        - "Max slope (smoothed)"

    Notes
    -----
    - Lag is defined as the first timepoint where OD exceeds
      `min(OD) * lag_threshold`.
    - Slope is computed using a numerical gradient.
    - Smoothed slope is computed after applying a Savitzky-Golay filter.
    """
    result = {}
    
    result["Lag (s)"] = result["Lag (hr)"] = None
    cutoff_ods = df["OD"] >= df["OD"].min() * lag_threshold
    if cutoff_ods.any():
        # We're creating a mask where all ODs larger than the cutoff will become NaN.
        # Therefore, the last valid index +1 will be the first index where the OD is
        # always larger than the cutoff.
        s = (~cutoff_ods).reset_index()["OD"]
        tt_od_index = s.where(s).last_valid_index()+1
        if tt_od_index < len(df):
            result["Lag (s)"] = df.iloc[tt_od_index]["Time (s)"]
            result["Lag (hr)"] = round(result["Lag (s)"] / 60 / 60, 2) # In hours, rounded, for convenience.

    result["Max OD"] = df["OD"].max()

    slope = np.gradient(df["OD"])
    result["Max slope"] = slope.max()

    # Because of potential "jumps" in the growth curve, we smooth it and recalculate the slope.
    # Parameters chosen empirically.
    from scipy.signal import savgol_filter
    try:
        smooth_od = savgol_filter(df["OD"], 10, 3)
    except:
        print(df.index.unique())
        raise
    result["Max slope (smoothed)"] = np.gradient(smooth_od).max()

    if plot:
        import matplotlib.pyplot as plt
        fig, axs = plt.subplots(nrows=3, figsize=(10, 10))
        axs[0].set_title(f"{df.index.unique().get_level_values(0)[0]} {df.index.unique().get_level_values(1)[0]}")
        axs[0].plot(range(len(df)), df["OD"], color="blue")
        axs[0].plot(range(len(df)), smooth_od, color="red")
        axs[1].plot(range(len(slope)), slope, color="blue")
        axs[1].plot(range(len(slope)), np.gradient(smooth_od), color="red")
        axs[2].plot(range(len(slope)), np.gradient(slope))
    
    return pd.Series(result)


def get_model_free_fitness_df(df: pd.DataFrame) -> pd.DataFrame:
    """Compute model-free fitness metrics for all growth curves.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing multiple growth curves.

    Returns
    -------
    pandas.DataFrame
        DataFrame where each row corresponds to a growth curve and columns
        contain fitness metrics.

    Notes
    -----
    Applies `get_model_free_fitness_single` to each group defined by the
    full index of `df`.
    """
    return df.groupby(df.index.names).apply(get_model_free_fitness_single)


################################################################################
# MICs
################################################################################

def get_mic_distributions(
    df: pd.DataFrame,
    timepoint: float = 24,
    inhibition: float = 0.5,
    timepoint_loc: Literal["left", "right"] = "left",
    target_od_method: Literal["median", "mean"] = "median",
    drug_ix_name: str = "FLC",
) -> pd.DataFrame:
    """Compute MIC distributions based on OD inhibition at a fixed timepoint.

    For each growth curve, the OD at a specified timepoint is compared to a
    target OD derived from the corresponding no-drug (e.g., FLC=0) condition.
    The MIC distribution is defined as the fraction of curves whose OD falls
    below this target at each drug concentration.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing growth curves. Must include:
        - Columns: "OD", "Time (s)"
        - A MultiIndex including `_Source` and a drug concentration level
          specified by `drug_ix_name`.
    timepoint : float, default: 24
        Time (in hours) at which OD values are evaluated.
    inhibition : float, default: 0.5
        Fractional inhibition used to define the target OD. The target is
        computed as:

            min_OD + (max_OD - min_OD) * (1 - inhibition)

        based on the no-drug condition at the `timepoint`.
    timepoint_loc : {"left", "right"}, default: "left"
        Strategy for selecting the timepoint:
        - "left": use the last timepoint <= specified timepoint
        - "right": use the first timepoint >= specified timepoint
    target_od_method : {"median", "mean"}, default: "median"
        Aggregation method used to compute the target OD across replicates.
    drug_ix_name : str, default: "FLC"
        Name of the index level corresponding to drug concentration.

    Returns
    -------
    pandas.DataFrame
        DataFrame indexed by all index levels except `_Source` and `drug_ix_name`,
        with columns corresponding to drug concentrations. Values represent the
        fraction of curves below the target OD (i.e., inhibited fraction).

    Notes
    -----
    - Timepoints are aligned by selecting a single measurement per curve;
      no interpolation is performed.

    Raises
    ------
    KeyError
        If the specified `drug_ix_name` level or required columns are missing.
    IndexError
        If no valid timepoint is found under the selected `timepoint_loc`.
    """

    # TODO: consider adding timepoint_loc="closest" that will take the closest timepoint to the specified one, regardless of whether it's on the left or right.
    # Also interpolated exactly at timepoint could be interesting, but would require more code to do properly
    # (e.g. checking that there are timepoints on both sides of the specified timepoint, and that they are not too far from it).

    # Timepoints to take for every curve:
    # We're setting the timepoint to be part of the index so that we can access the OD at that timepoint directly.
    if timepoint_loc == "left":
        timepoints_series = df[["Time (s)"]].groupby(df.index.names).agg( (lambda x: x[x <= timepoint*60*60].iloc[-1]) ).set_index("Time (s)", append=True)
    elif timepoint_loc == "right":
        timepoints_series = df[["Time (s)"]].groupby(df.index.names).agg( (lambda x: x[x >= timepoint*60*60].iloc[0]) ).set_index("Time (s)", append=True)

    # Get the ODs at the timepoints:
    ods_at_timepoint_series = df.set_index("Time (s)", append=True)["OD"].loc[timepoints_series.index].reset_index(level="Time (s)", drop=True)

    # Compute the target OD per every index based on its OD at the timepoint in drug=0.
    df_flc0 = df.xs(0, level=drug_ix_name).join(timepoints_series.reset_index("Time (s)"), rsuffix="_max")
    df_flc0 = df_flc0[ df_flc0["Time (s)"] <= df_flc0["Time (s)_max"] ]["OD"].droplevel(drug_ix_name)
    gb_obj = df_flc0.groupby(df_flc0.index.names)
    target_ods_series = (gb_obj.max() - gb_obj.min()) * (1-inhibition) + gb_obj.min()
    target_ods_series = target_ods_series.groupby(level=target_ods_series.index.names.difference(["_Source"]))
    if target_od_method == "median":
        target_ods_series = target_ods_series.median()
    elif target_od_method == "mean":
        target_ods_series = target_ods_series.mean()
    
    # Compare the OD at the timepoint to the target OD:
    df_both_ods = ods_at_timepoint_series.to_frame().join(target_ods_series, rsuffix="_target")
    is_below_target_od_series = df_both_ods["OD"] < df_both_ods["OD_target"]
    
    # Compute the fraction of curves that are below the target OD per every drug concentration:
    gb_obj = is_below_target_od_series.groupby(is_below_target_od_series.index.names.difference(["_Source"]))
    mics_df = (gb_obj.sum() / gb_obj.count()).unstack(level=drug_ix_name)

    return mics_df

def get_mic_ranges(
    df: pd.DataFrame,
    threshold: float = 1,
    agg_col: Optional[str] = None
) -> pd.DataFrame:
    """Convert MIC distributions into MIC ranges.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame where columns correspond to FLC concentrations and values
        represent MIC probabilities or scores.
    threshold : float, default: 1
        Minimum value required to consider a concentration as inhibitory.
    agg_col : str, optional
        Index level to aggregate over when computing MIC ranges.

    Returns
    -------
    pandas.DataFrame or pandas.Series
        If `agg_col` is None, returns MIC assignments per row.
        Otherwise, returns aggregated MIC ranges with columns:
        "from_mic" and "to_mic".

    Notes
    -----
    - MIC is defined as the lowest FLC where the value exceeds `threshold`.
    - Special cases:
        - Below lowest concentration → "<min"
        - Above highest concentration → ">max"
        - Zero concentration → "0?"
    """
    flcs = df.columns
    max_flc = max(flcs)
    min_flc = flcs[1]
    over_flc = max_flc+1
        
    result = df.apply(lambda x: x.idxmax() if x.max() >= threshold else over_flc, axis=1)

    def flc_to_mic(flc):
        return "0?" if flc == 0 else f"<{flc}" if flc == min_flc else f">{max_flc}" if flc == over_flc else flc
    
    if agg_col is not None:
        result = result.groupby(result.index.names.difference([agg_col])).\
            agg(
                from_mic=lambda x: flc_to_mic(x.min()),
                to_mic=lambda x: flc_to_mic(x.max())
            )
    else:
        result = result.apply(flc_to_mic)
    
    return result

try:
    from gooey import Gooey, GooeyParser
except ImportError:
    # If gooey isn't installed, provide no-op replacements so the script still runs.
    def Gooey(*args, **kwargs):
        def _decorator(func):
            return func
        return _decorator

    import argparse
    GooeyParser = argparse.ArgumentParser

# A kind of a fix for dark mode:
# item_default = {
#         'error_color': '#ea7878',
#         'label_color': '#ffffff',
#         'help_color': '#363636',
#         'full_width': False,
#         'validator': {
#             'type': 'local',
#             'test': 'lambda x: True',
#             'message': ''
#         },
#         'external_validator': {
#             'cmd': '',
#         }
#     }
# @Gooey(dump_build_config=True,
#         program_name="Widget Demo",
#         advanced=True,
#         auto_start=False,
#         body_bg_color='#262626',
#         header_bg_color='#262626',
#         footer_bg_color='#262626',
#         sidebar_bg_color='#262626',
#         )
@Gooey
def main():
    import argparse
    from openpyxl import load_workbook
    from growth_curves import read_od_sheet_with_plate_map

    # parser = argparse.ArgumentParser(description="Compute fitness from growth curve data.")
    parser = GooeyParser(description="Compute fitness from growth curve data.")
    parser.add_argument(
        'in_file',
        help='Excel file with the growth curves.',
        widget="FileChooser",
    )
    parser.add_argument(
        'input_format',
        # choices=['f200', 'spark', 'spark stacker']
        choices=['f200', 'spark_single', 'spark_stacked'],
        help="The machine which produced the growth curves. Note that if the SPARK only had one plate, you should specifiy 'spark_single', otherwise (if it ran in stacker mode) 'spark_stacked'.",
    )
    parser.add_argument(
        'plate_size',
        choices=[96, 384],
        type=int,
        help="The size of the plate used in the experiment.",
    )
    parser.add_argument(
        'out_file',
        help='Output file, CSV format.',
        widget="FileChooser",
    )

    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        '--sheets',
        help="Comma-separated list of sheets to process. If not given, all sheets will be processed. Mutually exclusive with '--exclude-sheets'.",
    )
    group.add_argument(
        '--exclude-sheets',
        help="Comma-separated list of sheets to exclude from processing. Mutually exclusive with '--sheets'.",
    )

    parser.add_argument(
        '--start-hour',
        help="The hour to start the growth curve from (i.e., drop all data before this timepoint). Default: 0.",
        type=float,
    )
    parser.add_argument(
        '--stop-hour',
        help="The hour to stop the growth curve at (counted from t=0, regardless of '--start-hour'). Default: the end of the growth curve.",
        type=float,
    )
    args = parser.parse_args()

    format = {'spark_single': 'spark', 'spark_stacked': 'spark stacker'}.get(args.input_format, args.input_format)

    wb = load_workbook(args.in_file)
    sheets = args.sheets.split(",") if args.sheets is not None else wb.sheetnames
    if args.exclude_sheets is not None:
        exclude_sheets = args.exclude_sheets.split(",")
        sheets = [s for s in sheets if s not in exclude_sheets]

    plate_indices = ["Row", "Column", "Cell", "Sheet"]
    plate_key_template = {}
    for row in "ABCDEFGH":
        for col in range(1, 13):
            plate_key_template[f"{row}{col}"] = (row, col, f"{row}{col}")

    dfs_to_concat = []
    for sheet_name in sheets:
        plate_key = {k: v + (sheet_name,) for k, v in plate_key_template.items()}
        df = read_od_sheet_with_plate_map(wb, sheet_name, plate_indices, plate_key, format, plate_type=args.plate_size, plate_name=sheet_name)
        
        start_hour = args.start_hour if args.start_hour is not None else 0
        stop_hour = args.stop_hour if args.stop_hour is not None else (df["Time (s)"].max() / 60 / 60)
        df = df[(df["Time (s)"] >= start_hour * 60 * 60) & (df["Time (s)"] <= stop_hour * 60 * 60)]
        dfs_to_concat.append(get_model_free_fitness_df(df))

    df = pd.concat(dfs_to_concat).sort_index(level=["Sheet", "Row", "Column"])
    df.to_csv(args.out_file)

if __name__ == "__main__":
    main()